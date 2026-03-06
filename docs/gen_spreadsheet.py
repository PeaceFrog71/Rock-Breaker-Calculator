"""Generate charge rate data collection spreadsheet for issue #278."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# -- Styles --
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
section_font = Font(bold=True, color="00D9FF", size=12)
section_fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
note_font = Font(italic=True, color="666666", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_input(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.fill = input_fill
    cell.alignment = center
    cell.border = thin_border


def style_calc(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.fill = calc_fill
    cell.alignment = center
    cell.border = thin_border


def style_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.alignment = center
    cell.border = thin_border


# ============================================
# Sheet 1: Data Collection
# ============================================
ws = wb.active
ws.title = "Data Collection"

ws.merge_cells("A1:J1")
ws["A1"] = "Charge Rate Data Collection - Issue #278"
ws["A1"].font = Font(bold=True, color="00D9FF", size=14)
ws["A1"].fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:J2")
ws["A2"] = (
    "Yellow cells = fill in during testing. Green cells = auto-calculated. "
    "Keep rock properties consistent within each test group."
)
ws["A2"].font = note_font
ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

# Section header
ws.merge_cells("A4:J4")
ws["A4"] = "ROCK PROPERTIES (keep consistent per test group)"
ws["A4"].font = section_font
ws["A4"].fill = section_fill

row = 5
headers = [
    "Test #", "Rock Mass", "Rock Resistance", "Laser Head",
    "Module 1", "Module 2", "Module 3", "Gadget",
    "Time to Fill (sec)", "Notes",
]
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header(ws, row, len(headers))

# Pre-populated test rows
tests = [
    (1, "", "", "Lancet MH1", "---", "---", "---", "---", "", "BASELINE: No mods, no gadgets"),
    (2, "", "", "Hofstede-S1", "---", "---", "---", "---", "", "Baseline: Different laser head"),
    (3, "", "", "Helix I", "---", "---", "---", "---", "", "Baseline: Different laser head"),
    (4, "", "", "Arbor MH1", "---", "---", "---", "---", "", "Baseline: Different laser head"),
    (5, "", "", "Impact I", "---", "---", "---", "---", "", "Baseline: Different laser head"),
    (6, "", "", "Klein-S1", "---", "---", "---", "---", "", "Baseline: Different laser head"),
    (7, "", "", "", "", "", "", "", "", ""),
    (8, "", "", "Lancet MH1", "Torrent", "---", "---", "---", "", "Single mod: +30% charge rate"),
    (9, "", "", "Lancet MH1", "Torrent II", "---", "---", "---", "", "Single mod: +35% charge rate"),
    (10, "", "", "Lancet MH1", "Torrent III", "---", "---", "---", "", "Single mod: +40% charge rate"),
    (11, "", "", "Lancet MH1", "Torpid", "---", "---", "---", "", "Single mod: +60% charge rate (active)"),
    (12, "", "", "", "", "", "", "", "", ""),
    (13, "", "", "Lancet MH1", "---", "---", "---", "Stalwart", "", "Gadget only: +50% charge rate"),
    (14, "", "", "Lancet MH1", "---", "---", "---", "Waveshift", "", "Gadget only: -30% charge rate"),
    (15, "", "", "", "", "", "", "", "", ""),
    (16, "", "", "Hofstede-S1", "Torrent III", "---", "---", "Stalwart", "", "FASTEST: High CR laser + mod + gadget"),
    (17, "", "", "Lancet MH1", "Torrent III", "---", "---", "Stalwart", "", "FASTEST: Lancet variant"),
    (18, "", "", "", "", "", "", "", "", ""),
    (19, "", "", "Arbor MH1", "Focus", "---", "---", "Waveshift", "", "SLOWEST: No CR bonus + negative gadget"),
    (20, "", "", "", "", "", "", "", "", ""),
]

for i, test in enumerate(tests):
    r = 6 + i
    for col, val in enumerate(test, 1):
        ws.cell(row=r, column=col, value=val)
        if col in (2, 3, 9):
            style_input(ws, r, col)
        else:
            style_cell(ws, r, col)

# Blank rows for additional tests
for i in range(20):
    r = 26 + i
    ws.cell(row=r, column=1, value=21 + i)
    for col in range(1, 11):
        if col in (2, 3, 9):
            style_input(ws, r, col)
        else:
            style_cell(ws, r, col)

col_widths = [8, 12, 14, 16, 14, 14, 14, 14, 16, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w


# ============================================
# Sheet 2: Equipment Reference
# ============================================
ws2 = wb.create_sheet("Equipment Reference")

ws2.merge_cells("A1:D1")
ws2["A1"] = "Charge Rate Modifiers - From Equipment Database"
ws2["A1"].font = Font(bold=True, color="00D9FF", size=14)
ws2["A1"].fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")

# Laser heads
r = 3
ws2.cell(row=r, column=1, value="LASER HEADS")
ws2.cell(row=r, column=1).font = section_font
ws2.cell(row=r, column=1).fill = section_fill

r = 4
for col, h in enumerate(["Laser Head", "Charge Rate Modifier", "Effect", "Size"], 1):
    ws2.cell(row=r, column=col, value=h)
style_header(ws2, r, 4)

lasers = [
    ("Lancet MH1", 1.4, "+40%", "S1"),
    ("Lancet MH2", 1.4, "+40%", "S2"),
    ("Hofstede-S1", 1.2, "+20%", "S1"),
    ("Hofstede-S2", 1.2, "+20%", "S2"),
    ("Pitman", 0.6, "-40%", "S0"),
    ("Arbor MH1", 1.0, "None", "S1"),
    ("Arbor MH2", 1.0, "None", "S2"),
    ("Helix I", 1.0, "None", "S1"),
    ("Helix II", 1.0, "None", "S2"),
    ("Impact I", 0.6, "-40%", "S1"),
    ("Impact II", 0.6, "-40%", "S2"),
    ("Klein-S1", 1.0, "None", "S1"),
    ("Klein-S2", 1.0, "None", "S2"),
]
for i, (name, mod, effect, size) in enumerate(lasers):
    r = 5 + i
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=mod)
    ws2.cell(row=r, column=3, value=effect)
    ws2.cell(row=r, column=4, value=size)
    for c in range(1, 5):
        style_cell(ws2, r, c)

# Modules
r = 20
ws2.cell(row=r, column=1, value="MODULES")
ws2.cell(row=r, column=1).font = section_font
ws2.cell(row=r, column=1).fill = section_fill

r = 21
for col, h in enumerate(["Module", "Charge Rate Modifier", "Effect", "Type"], 1):
    ws2.cell(row=r, column=col, value=h)
style_header(ws2, r, 4)

modules = [
    ("Torrent", 1.3, "+30%", "Passive"),
    ("Torrent II", 1.35, "+35%", "Passive"),
    ("Torrent III", 1.4, "+40%", "Passive"),
    ("Torpid", 1.6, "+60%", "Active"),
    ("Vaux", 0.8, "-20%", "Passive"),
    ("Vaux-C2", 0.85, "-15%", "Passive"),
    ("Vaux-C3", 0.95, "-5%", "Passive"),
]
for i, (name, mod, effect, typ) in enumerate(modules):
    r = 22 + i
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=mod)
    ws2.cell(row=r, column=3, value=effect)
    ws2.cell(row=r, column=4, value=typ)
    for c in range(1, 5):
        style_cell(ws2, r, c)

# Gadgets
r = 31
ws2.cell(row=r, column=1, value="GADGETS")
ws2.cell(row=r, column=1).font = section_font
ws2.cell(row=r, column=1).fill = section_fill

r = 32
for col, h in enumerate(["Gadget", "Charge Rate Modifier", "Effect", ""], 1):
    ws2.cell(row=r, column=col, value=h)
style_header(ws2, r, 4)

gadgets = [
    ("Stalwart", 1.5, "+50%"),
    ("Waveshift", 0.7, "-30%"),
]
for i, (name, mod, effect) in enumerate(gadgets):
    r = 33 + i
    ws2.cell(row=r, column=1, value=name)
    ws2.cell(row=r, column=2, value=mod)
    ws2.cell(row=r, column=3, value=effect)
    for c in range(1, 4):
        style_cell(ws2, r, c)

for i, w in enumerate([18, 22, 16, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ============================================
# Sheet 3: Analysis
# ============================================
ws3 = wb.create_sheet("Analysis")

ws3.merge_cells("A1:F1")
ws3["A1"] = "Analysis - Fill in after data collection"
ws3["A1"].font = Font(bold=True, color="00D9FF", size=14)
ws3["A1"].fill = PatternFill(start_color="0A0E27", end_color="0A0E27", fill_type="solid")

ws3.merge_cells("A2:F2")
ws3["A2"] = "Use this sheet to derive the base charge rate and validate modifier math"
ws3["A2"].font = note_font
ws3["A2"].alignment = Alignment(horizontal="center")

r = 4
headers = [
    "Test #", "Combined CR Modifier", "Measured Time (sec)",
    "Charge Rate (%/sec)", "Expected Ratio vs Baseline", "Actual Ratio vs Baseline",
]
for col, h in enumerate(headers, 1):
    ws3.cell(row=r, column=col, value=h)
style_header(ws3, r, len(headers))

for i in range(30):
    r = 5 + i
    ws3.cell(row=r, column=1, value=i + 1)
    style_cell(ws3, r, 1)
    for c in range(2, 7):
        style_cell(ws3, r, c)
    # Col B, C = manual input
    style_input(ws3, r, 2)
    style_input(ws3, r, 3)
    # Col D: Charge rate = 100 / time
    ws3.cell(row=r, column=4).value = f"=IF(C{r}<>\"\", 100/C{r}, \"\")"
    style_calc(ws3, r, 4)
    # Col E: Expected ratio = combined modifier (manual)
    style_input(ws3, r, 5)
    # Col F: Actual ratio vs baseline (row 5 = test 1)
    ws3.cell(row=r, column=6).value = f"=IF(AND(D{r}<>\"\",D$5<>\"\"), D{r}/D$5, \"\")"
    style_calc(ws3, r, 6)

for i, w in enumerate([8, 22, 18, 18, 22, 22], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# Save
path = (
    r"d:\OneDrive\Documents\Business\PeaceFrogGaming"
    r"\Applications\RockBreakerCalculator\docs\charge-rate-data-collection.xlsx"
)
wb.save(path)
print(f"Saved to {path}")
